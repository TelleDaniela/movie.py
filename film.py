from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import random

def moveUp(ws, row):
    max_row = ws.max_row
    row = int(row)
    max_row = int(max_row)
    for rovwal in range(row, max_row):
        for colval in range(1,7):
            if colval == 1:
                if ws.cell(rovwal+1, colval).value:
                    ws.cell(rovwal, colval).value = ws.cell(rovwal+1, colval).value-1
                    ws.cell(rovwal+1, colval).value = None
            else:
                ws.cell(rovwal, colval).value = ws.cell(rovwal+1, colval).value
                ws.cell(rovwal+1, colval).value = None
    return max_row-1



def addMovie(ws,file_path):
    name = input('Please enter the name of the movie \n')
    if not name:
        print('An error has occured, name is neccesary')
        return
    description = input('Please enter the description of the movie (Optional) \n')
    watched = input('Have you already watched the movie? Yes/No')
    if watched.lower() == 'Yes'.lower():
        watched = 'Yes'
        rating = input('Please rate the movie on a scale from 0 to 10 \n')
        if int(rating) > 10 or int(rating) < 0:
            print('Error, rating should be in interval between 0 and 10')
            return
    elif watched.lower() == 'No'.lower():
        watched = 'No'
    else:
        print('Error, please enter Yes or No')
        return
    duration = input('Please enter the duration of the movie(hh:mm:ss)')
    a = duration.split(':')
    if int(a[1]) > 99 or int(a[1]) < 0:
        print('wrong hour count')
        return
    if int(a[2]) > 59 or int(a[2]) < 0:
        print('wrong minute count')
        return
    if int(a[3]) > 60 or int(a[3]) < 0:
        print('wrong second count')
        return
    row = ws.max_row
    for rowval in range(1, row):
        if not ws.cell(rowval,1).value:
            ws['a'+str(rowval)] = rowval
            ws['a'+str(rowval)].alignment = Alignment(horizontal='left',vertical='center')
            ws['b'+str(rowval)] = name
            ws.column_dimensions['B'].width = 15
            ws['b'+str(rowval)].alignment = Alignment(horizontal='center',vertical='center')
            ws.column_dimensions['C'].width = 30
            ws['c'+str(rowval)] = description
            ws['d'+str(rowval)] = watched
            ws.column_dimensions['D'].width = 15
            ws['d'+str(rowval)].alignment = Alignment(horizontal='left',vertical='center')
            ws['e'+str(rowval)] = rating
            ws['e'+str(rowval)].alignment = Alignment(horizontal='center',vertical='center')
            ws['f'+str(rowval)].number_format = 'hh:mm:ss'
            ws['f'+str(rowval)] = duration
            wb.save(file_path)
            break 
    
    return

def delMovie(ws,file_path):
    mode = input('Enter 1 if you want to find and delete movie by id, or enter 2 if you want to find and delete movie by name')
    max_row = ws.max_row
    voidrow = -1
    if mode == '1':
        id = input('Enter id')
        for row in range(2,max_row):
            voidrow = row
            if ws['a'+str(row)].value == int(id):
                for rowval in ws['a'+str(row):'f'+str(row)]:
                    for cell in rowval:
                        cell.value = None
                moveUp(ws, voidrow)
                break
    elif mode == '2':
        name = input('Enter name')
        for row in range(2,max_row):
            voidrow = row
            if ws['b'+str(row)].value == name:
                for rowval in ws['a'+str(row):'f'+str(row)]:
                    for cell in rowval:
                        cell.value = None
                moveUp(ws, voidrow)
    else:
        print('Error, enter 1 or 2')
        return
    wb.save(file_path)

def editMovie(ws,file_path):
    mode = input('Enter 1 if you want to find and edit movie by id, or enter 2 if you want to find and edit movie by name')
    max_row = ws.max_row
    if mode == '1':
        id = input('Enter id')
        for row in range(2,max_row):
            if ws['a'+str(row)].value == int(id):
                name = input('Please enter the name of the movie \n')
                if not name:
                    print('An error has occured, name is neccesary')
                    return
                description = input('Please enter the description of the movie (Optional) \n')
                watched = input('Have you already watched the movie? Yes/No')
                if watched.lower() == 'Yes'.lower():
                    watched = 'Yes'
                    rating = input('Please rate the movie on a scale from 0 to 10 \n')
                    if int(rating) > 10 or int(rating) < 0:
                        print('Error, rating should be in interval between 0 and 10')
                        return
                elif watched.lower() == 'No'.lower():
                    watched = 'No'
                else:
                    print('Error, please enter Yes or No')
                    return
                duration = input('Please enter the duration of the movie(hh:mm:ss)')
                if int(a[1]) > 99 or int(a[1]) < 0:
                    print('wrong hour count')
                    return
                if int(a[2]) > 59 or int(a[2]) < 0:
                    print('wrong minute count')
                    return
                if int(a[3]) > 60 or int(a[3]) < 0:
                    print('wrong second count')
                    return
                id = int(id)
                ws['a'+str(id+1)] = id
                ws['a'+str(id+1)].alignment = Alignment(horizontal='left',vertical='center')
                ws['b'+str(id+1)] = name
                ws.column_dimensions['B'].width = 15
                ws['b'+str(id+1)].alignment = Alignment(horizontal='center',vertical='center')
                ws.column_dimensions['C'].width = 30
                ws['c'+str(id+1)] = description
                ws['d'+str(id+1)] = watched
                ws.column_dimensions['D'].width = 15
                ws['d'+str(id+1)].alignment = Alignment(horizontal='left',vertical='center')
                ws['e'+str(id+1)] = rating
                ws['e'+str(id+1)].alignment = Alignment(horizontal='center',vertical='center')
                ws['f'+str(id+1)].number_format = 'hh:mm:ss'
                ws['f'+str(id+1)] = duration
            break 
                            
    elif mode == '2':
        name = input('Enter name')
        for row in range(2,max_row):
            if ws['b'+str(row)].value == name:
                name = input('Please enter the name of the movie \n')
                if not name:
                    print('An error has occured, name is neccesary')
                    return
                description = input('Please enter the description of the movie (Optional) \n')
                watched = input('Have you already watched the movie? Yes/No')
                if watched.lower() == 'Yes'.lower():
                    watched = 'Yes'
                    rating = input('Please rate the movie on a scale from 0 to 10 \n')
                    if int(rating) > 10 or int(rating) < 0:
                        print('Error, rating should be in interval between 0 and 10')
                        return
                elif watched.lower() == 'No'.lower():
                    watched = 'No'
                else:
                    print('Error, please enter Yes or No')
                    return
                duration = input('Please enter the duration of the movie(hh:mm:ss)')
                if int(a[1]) > 99 or int(a[1]) < 0:
                    print('wrong hour count')
                    return
                if int(a[2]) > 59 or int(a[2]) < 0:
                    print('wrong minute count')
                    return
                if int(a[3]) > 60 or int(a[3]) < 0:
                    print('wrong second count')
                    return
                
                ws['b'+str(row)] = name
                ws.column_dimensions['B'].width = 15
                ws['b'+str(row)].alignment = Alignment(horizontal='center',vertical='center')
                ws.column_dimensions['C'].width = 30
                ws['c'+str(row)] = description
                ws['d'+str(row)] = watched
                ws.column_dimensions['D'].width = 15
                ws['d'+str(row)].alignment = Alignment(horizontal='left',vertical='center')
                ws['e'+str(row)] = rating
                ws['e'+str(row)].alignment = Alignment(horizontal='center',vertical='center')
                ws['f'+str(row)].number_format = 'hh:mm:ss'
                ws['f'+str(row)] = duration
            break
    else:
        print('Error, enter 1 or 2')
        return
    wb.save(file_path)

def randMovie(ws):
    max_rows = ws.max_row
    rand = random.randint(1,max_rows)
    print(ws.cell(rand,1).value,ws.cell(rand,2).value,ws.cell(rand,3).value,ws.cell(rand,4).value,ws.cell(rand,5).value,ws.cell(rand,6).value)

def printMovies(ws):
    max = ws.max_row
    for rand in range(1, max):
        print(ws.cell(rand,1).value,ws.cell(rand,2).value,ws.cell(rand,3).value,ws.cell(rand,4).value,ws.cell(rand,5).value,ws.cell(rand,6).value)


while True:
    path = input('Enter a path to your table, or enter 1 to open last used table, or enter 2 to create table \n')
    if path == '1':
        try:
            f =open('C:/path_to_table.txt','r')
            path = f.readline()
        except:
            print('An error occured, no recently opened table was found')
            continue
        if not path:
            print('An error occured, no recently opened table was found')
            continue
        else:
            
            wb=load_workbook(path)
        break
    elif path == '2':
        name = input('Enter a name for your table')
        table_path = input('Enter a path to folder you want to create a table in')
        wb = Workbook()
        try:
            wb.save(table_path + '/' + name+'.xlsx')
        except:
            print('Error, wrong path')
            continue
        ws = wb.worksheets[0]
        with open('C:/path_to_table.txt','w') as f:
            f.write(table_path + '/' + name+'.xlsx')
            f.close()
        path = table_path + '/' + name+'.xlsx'
        ws['A1'] = 'ID'
        ws['B1'] = 'Movie Name'
        ws['C1'] = 'Description'
        ws['D1'] = 'Watched Y/N'
        ws['E1'] = 'Rating'
        ws['F1'] = 'Duration'
        break
    else:
        try:
            wb=load_workbook(path)
        except:
            print('An error occured, try diffrent path')
            continue
        with open('C:/path_to_table.txt','w') as f:
            f.write(path)
            f.close()
        break
ws = wb.worksheets[0]

while True:
    print('a)Add a movie')
    print('b)Delete a movie')
    print('c)Edit a movie data')
    print('d)Pick a random movie to watch')
    print('e)Print all movies to the console')
    print('g)Close program')

    a = input()
    if a == 'a':
        addMovie(ws, path)
    elif a == 'b':
        delMovie(ws,path)
    elif a == 'c':
        editMovie(ws,path)
    elif a == 'd':
        randMovie(ws)
    elif a == 'e':
        printMovies(ws)
    elif a == 'g':
        print('Goodbye!')
        break
    else:
        print('Command not supported, try another')


